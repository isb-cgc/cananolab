"""
Copyright 2020-2024, Institute for Systems Biology

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

   http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

from openpyxl import load_workbook
import sys
import json

json_dict = {
  "data": {
    "type": "dois",
    "attributes": {
      "prefix": None,
      "creators": [],
      "titles": [
        {
          "lang": None,
          "title": None,
          "titleType": None
        }
      ],
      "language": "en",
      "publisher": "NCI caNanoLab",
      "publicationYear": None,
      "contributors": [],
      "types": {
        "resourceTypeGeneral": None
      },
      "url": None,
      "version": None,
      "descriptions": [
        {
          "lang": None,
          "description": None,
          "descriptionType": None
        }
      ]
    }
  }
}

def main(args):

    if len(args) != 4:
        print(" Usage : %s <prod | test> <excel_file> <json_output>" % args[0])
        sys.exit(-1)

    tier = args[1]
    excel_file = args[2]
    json_output = args[3]

    if tier == "prod":
        doi_prefix = "10.17917"
    else:
        doi_prefix = "10.24368"

    json_dict['data']['attributes']['prefix'] = doi_prefix

    wb = load_workbook(excel_file, data_only = True)

    #
    # Process Creator Sheet
    #

    creators_sheet = wb["Creators"]

    heading_row = True
    creators = json_dict['data']['attributes']['creators']

    for row in creators_sheet.iter_rows():
        if heading_row:
            heading_row = False
            continue
        creator = {"nameType": "Personal"}
        if row[0].value is not None:
            creator['givenName'] = row[0].value
            creator['familyName'] = row[1].value
            creator['name'] = row[2].value
            creators.append(creator)
        else:
            break

    #
    # Process Contributors Sheet
    #

    contribs_sheet = wb["Contributors"]
    contribs = json_dict['data']['attributes']['contributors']

    heading_row = True
    for row in contribs_sheet.iter_rows():
        if heading_row:
            heading_row = False
            continue
        contrib = {"nameType": "Personal"}
        if row[0].value is not None:
            contrib['givenName'] = row[0].value
            contrib['familyName'] = row[1].value
            contrib['name'] = row[2].value
            contrib['contributorType'] = row[3].value
            contribs.append(contrib)
        else:
            break

    #
    # Process "All Other Fields" Sheet
    #

    fields = wb["All Other Fields"]
    attributes = json_dict['data']['attributes']

    heading_row = True
    for row in fields.iter_rows():
        if heading_row:
            heading_row = False
            continue

        if row[0].value is not None:
            attributes['titles'][0]['title'] = row[0].value
            attributes['publicationYear'] = row[1].value
            attributes['version'] = row[2].value
            attributes['types']['resourceTypeGeneral'] = row[3].value
            attributes['url'] = row[4].value
            attributes['descriptions'][0]['descriptionType'] = row[5].value
            attributes['descriptions'][0]['description'] = row[6].value
        else:
            break

    #
    # Write out the JSON file
    #

    json_object = json.dumps(json_dict, indent=4)
    print(json_object)
    print(json_output)
    with open(json_output, "w") as outfile:
        outfile.write(json_object)

    return

if __name__ == '__main__':
    main(sys.argv)